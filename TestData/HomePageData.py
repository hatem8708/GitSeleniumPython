
import openpyxl

class HomePageData:

    test_HomePage_data =[{"firstname":"Hatem", "email":"name@email.com", "gender":"Male"}, {"firstname":"Hatem","email":"name@gmail.com", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):

        book = openpyxl.load_workbook("C:\\Users\\Admin\\Desktop\\PythonDemo.xlsx")
        sheet = book.active
        Dic = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "test_case_name":
                for j in range(2, sheet.max_column + 1):
                    Dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dic]